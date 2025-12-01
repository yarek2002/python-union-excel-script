import os
import traceback
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from collections import Counter, defaultdict

def is_numeric(s):
    if pd.isna(s):
        return False
    try:
        float(str(s).replace(',', '.'))
        return True
    except:
        return False

def unique_within_file(headers):
    count = Counter(headers)
    version = defaultdict(int)
    result = []
    for h in headers:
        if count[h] > 1:
            version[h] += 1
            result.append(f"{h}-{version[h]}")
        else:
            result.append(h)
    return [h.replace("Дата", f"Дата-{i+1}") if h.startswith("Дата") else h for i, h in enumerate(result)]

def find_header_info(file_path):
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and str(v).startswith("№"):
                headers = []
                start_col = c
                cc = c
                while cc <= ws.max_column:
                    cell = ws.cell(r, cc)
                    if cell.value is None:
                        break
                    headers.append(str(cell.value))
                    cc += 1
                return r - 1, start_col - 1, headers
    return 0, 0, []

def get_union_headers(folder_path):
    return ['Файл', '№', 'Запрос от', 'Комментарий от', 'Документ', 'Раздел', 'Лист',
            'Дата-1', 'Дата-2', 'Комментарий Заказчика', 'Ответ Проектной Организации',
            'Текущий статус','Статус (примечание)','Количество итераций']

def extract_file_data(file_path):
    raw = pd.read_excel(file_path, header=None, engine="openpyxl", dtype=str)
    header_start, start_col, headers = find_header_info(file_path)
    if not headers:
        return None

    headers_unique = unique_within_file(headers)
    body = raw.iloc[header_start + 1:, start_col: start_col + len(headers_unique)].copy()

    if body.empty:
        print(f"Нет строк под шапкой: {os.path.basename(file_path)}")
        return None

    body.columns = headers_unique
    body = body.dropna(how="all").reset_index(drop=True)

    if body.empty:
        print(f"Все строки пустые после dropna: {os.path.basename(file_path)}")
        return None

    records = []
    for i in range(len(body)):
        val = body.iloc[i, 0]

        if is_numeric(val) and str(val).strip() != "":
            row = body.iloc[i]

            record = {
                "Файл": os.path.basename(file_path),
                "№": str(val).strip()
            }

            record["Запрос от"] = row.get("Запрос от-1", row.get("Запрос от", pd.NA))
            record["Комментарий от"] = row.get("Комментарий от-1", row.get("Комментарий от", pd.NA))
            record["Документ"] = row.get("№ документа-1", row.get("Название документа-1",
                                row.get("№ документа", row.get("Название документа", pd.NA)))
            )
            record["Раздел"] = row.get("Раздел-1", row.get("Раздел", pd.NA))
            record["Лист"] = row.get("Лист-1", row.get("Лист", pd.NA))

            date_vals = [row[c] for c in body.columns if c.startswith("Дата") and not pd.isna(row[c])]
            if len(date_vals) == 1:
                d = pd.to_datetime(date_vals[0], errors='coerce')
                if not pd.isna(d):
                    record["Дата-1"] = d.strftime("%d-%m-%Y")
            elif len(date_vals) > 1:
                d1 = pd.to_datetime(date_vals[0], errors='coerce')
                d2 = pd.to_datetime(date_vals[-1], errors='coerce')
                if not pd.isna(d1): record["Дата-1"] = d1.strftime("%d-%m-%Y")
                if not pd.isna(d2): record["Дата-2"] = d2.strftime("%d-%m-%Y")

            cust_vals = [row[c] for c in body.columns if "Комментарий Заказчика" in c and not pd.isna(row[c])]
            if cust_vals:
                cust_list = [str(v).strip() for v in cust_vals if str(v).strip().lower() not in ["nan","none",""]]
                if cust_list:
                    record["Комментарий Заказчика"] = "\n".join(f"{idx+1}) {c}" for idx, c in enumerate(cust_list))

            ans_vals = [row[c] for c in body.columns if "Ответ Проектной Организации" in c and not pd.isna(row[c])]
            if ans_vals:
                ans_list = [str(v).strip() for v in ans_vals if str(v).strip().lower() not in ["nan","none",""]]
                if ans_list:
                    record["Ответ Проектной Организации"] = "\n".join(f"{idx+1}) {a}" for idx, a in enumerate(ans_list))

                        #  Определяем последний заполненный столбец из нужных групп
            check_cols = []
            for c in body.columns:
                if ("Статус" in c) or ("Ответ" in c) or ("Проектной Организации" in c) or ("Комментарий Заказчика" in c):
                    check_cols.append(c)

            last_filled_col = None
            for c in reversed(check_cols):  # идём с конца вправо
                if not pd.isna(row[c]) and str(row[c]).strip() != "":
                    last_filled_col = c
                    break

            #  Заполняем "Статус (примечание)" только если последний непустой — это статус
            if last_filled_col and "Статус" in last_filled_col:
                record["Статус (примечание)"] = str(row[last_filled_col]).strip()
            else:
                record["Статус (примечание)"] = ""  # пусто, если последний был ответом или комментарием

                        

                        # 1. Списки колонок по группам (в порядке как они есть в файле Excel)
            status_cols  = [c for c in body.columns if "Статус" in c]
            answer_cols  = [c for c in body.columns if "Ответ Проектной Организации" in c]
            comment_cols = [c for c in body.columns if "Комментарий Заказчика" in c]

            # 2. Собираем все 3 группы в порядке файла, без пересборки вручную
            target_cols = [c for c in body.columns if c in (comment_cols + answer_cols + status_cols)]

            # 3. Ищем самый правый заполненный столбец в ЭТОЙ строке
            last_filled_col = None
            last_filled_val = ""

            for col in reversed(target_cols):
                val = row[col]
                if not pd.isna(val) and str(val).strip() != "":
                    last_filled_col = col
                    last_filled_val = str(val).strip()
                    break

            # 4. Заполняем "Статус (примечание)" только если последний заполненный столбец был статусом
            if last_filled_col and last_filled_col in status_cols:
                record["Статус (примечание)"] = last_filled_val
            else:
                record["Статус (примечание)"] = ""

            # 5. Определяем "Текущий статус" по тому, к какой группе относился последний заполненный столбец
            if last_filled_col in comment_cols:
                record["Текущий статус"] = "Не снято"
            elif last_filled_col in answer_cols:
                record["Текущий статус"] = "Отработано"
            elif last_filled_col in status_cols:
                record["Текущий статус"] = "Исправлено"

            # 6. Считаем итерации ТОЛЬКО по заполненным Комментарий Заказчика В ЭТОЙ строке
            iteration_filled = 0
            for col in comment_cols:
                v = row[col]
                if not pd.isna(v) and str(v).strip().lower() not in ["nan", "none", ""]:
                    iteration_filled += 1

            record["Количество итераций"] = iteration_filled




            records.append(record)

    if not records:
        print(f"В файле {os.path.basename(file_path)} не найдено строк с числовым №.")
        return None

    return records

def process_folder(folder_path, output_file):
    files = [f for f in os.listdir(folder_path) if f.endswith(".xlsx") and not f.startswith("объединенный файл")]
    records = []
    for name in files:
        path = os.path.join(folder_path, name)
        r = extract_file_data(path)
        if r:
            records.extend(r)

    merged = pd.DataFrame(records, columns=get_union_headers(folder_path))

    merged.to_excel(output_file, index=False)
    return output_file

if __name__ == "__main__":
    try:
        folder = os.getcwd()
        out = f"объединенный файл {datetime.now().strftime('%Y-%m-%d')}.xlsx"
        res = process_folder(folder, out)
        print("Готово:", res)
    except Exception:
        tb = traceback.format_exc()
        print("Ошибка:\n", tb)
        with open("error_log.txt", "w", encoding="utf-8") as f:
            f.write(tb)
    finally:
        os.system("pause")
